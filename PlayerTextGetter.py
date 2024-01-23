import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
import random
from datetime import date, datetime, timedelta
import csv
import pandas as pd
from io import StringIO
import os
import tensorflow as tf
from tensorflow import _keras_module
from keras.models import load_model
import cv2
import imghdr
import numpy as np
from matplotlib import pyplot as plt
from keras.models import Sequential
from keras.layers import Conv2D, MaxPooling2D, Dense, Flatten, Dropout
from keras.metrics import Precision, Recall, BinaryAccuracy


nationData = pd.read_csv('nations.csv')
# Convert the DataFrame to a Dictionary
temp_nations = nationData.to_dict(orient='list')
nations_Dict = dict(zip(temp_nations['Nation Name'], temp_nations['Nation ID']))

playerNames = pd.read_csv('dcplayernames.csv')
temp_Names = playerNames.to_dict(orient='list')
names_Dict = dict(zip(temp_Names['name'], temp_Names['nameid']))


used_numbers = []



def generate_random_number(used_numbers):
    """Generate a random number between 1 and 99 that is not in the used_numbers set."""
    while True:
        random_number = random.randint(1, 99)
        if random_number not in used_numbers:
            used_numbers.append(random_number)
            return random_number



def dateIDGenerator(year, month, day):
    l_date = date(year, month, day)
    f_date = date(1753, 1, 1)
    delta = l_date - f_date
    id = 62171 + int(delta.days)
    return id



def ageGetter(id):
    base_date = date(1753, 1, 1)
    delta_days = id - 62171
    l_date = base_date + timedelta(days=delta_days)
    
    today = date.today()
    years_passed = today.year - l_date.year - ((today.month, today.day) < (l_date.month, l_date.day))
    
    return int(years_passed)




def attackerEstimation(age, value, avgLeagueSquadValue):
    if avgLeagueSquadValue >= 150:
        valueModifier = (0.1*value)
        qualityModifier = (0.005*551)
        baseRating = 70
        ageModifier = (21-age)*0.9
        rating = round((valueModifier+qualityModifier+baseRating)-ageModifier)
        return rating
        #print("Player rating is: " + str(rating))
    else:
        valueModifier = (0.5*value)
        qualityModifier = (0.005*avgLeagueSquadValue)
        baseRating = 65
        ageModifier = (21-age)
        rating = round((valueModifier+qualityModifier+baseRating)-ageModifier)
        return rating
        #print("Player rating is: " + str(rating))






def download_image(urltoDownload, filename):
        response = requests.get(urltoDownload)
        with open(filename, 'wb') as f:
            f.write(response.content)

























def predictHairstyle(urltoDownload):
    # Download the image from the URL
    img_filename = 'downloaded_image.webp'
    download_image(urltoDownload, img_filename)

    img = cv2.imread(img_filename)
    resize = tf.image.resize(img, (256,256))

    new_model = load_model(os.path.join('models', 'hairStyleClassifier.h5'))
    predVals = new_model.predict(np.expand_dims(resize/255, 0))
    maxVal = np.max(predVals[0])

    # Access each probability individually and convert to regular float
    c1 = float(predVals[0][0])
    c2 = float(predVals[0][1])
    c3 = float(predVals[0][2])
    c4 = float(predVals[0][3])
    c5 = float(predVals[0][4])
    c6 = float(predVals[0][5])
    c7 = float(predVals[0][6])
    Codes = [c1, c2, c3, c4, c5, c6, c7]

    for code in Codes:
        if code == c1 and code == maxVal:
            return 0
        elif code == c2 and code == maxVal:
            return 14
        elif code == c3 and code == maxVal:
            return 18
        elif code == c4 and code == maxVal:
            return 24
        elif code == c5 and code == maxVal:
            return 25
        elif code == c6 and code == maxVal:
            return 82
        elif code == c7 and code == maxVal:
            return 275






def predictFacialHair(urltoDownload):
    # Download the image from the URL
    img_filename = 'downloaded_image.webp'
    download_image(urltoDownload, img_filename)

    img = cv2.imread(img_filename)
    resize = tf.image.resize(img, (256,256))

    new_model = load_model(os.path.join('models', 'facialHairClassifier.h5'))
    predVal = new_model.predict(np.expand_dims(resize/255, 0))
    maxVal = np.max(predVal[0])

    # Access each probability individually and convert to regular float
    c1 = float(predVal[0][0])
    c2 = float(predVal[0][1])
    c3 = float(predVal[0][2])
    c4 = float(predVal[0][3])
    c5 = float(predVal[0][4])
    Codes = [c1, c2, c3, c4, c5]

    for code in Codes:
        if code == maxVal and code == c1:
            return 0
        elif code == maxVal and code == c2:
            return 4
        elif code == maxVal and code == c3:
            return 5
        elif code == maxVal and code == c4:
            return 8
        elif code == maxVal and code == c5:
            return 7




def predictSkinTone(urltoDownload):
    # Download the image from the URL
    img_filename = 'downloaded_image.webp'
    download_image(urltoDownload, img_filename)

    img = cv2.imread(img_filename)
    resize = tf.image.resize(img, (256,256))

    new_model = load_model(os.path.join('models', 'skinToneClassifier.h5'))
    predVal = new_model.predict(np.expand_dims(resize/255, 0))
    maxVal = np.max(predVal[0])

    # Access each probability individually and convert to regular float
    c1 = float(predVal[0][0])
    c2 = float(predVal[0][1])
    c3 = float(predVal[0][2])
    c4 = float(predVal[0][3])
    c5 = float(predVal[0][4])
    c6 = float(predVal[0][5])
    skinToneCodes = [c1, c2, c3, c4, c5, c6]

    for code in skinToneCodes:
        if code == maxVal and code == c1:
            return 1
        elif code == maxVal and code == c2:
            return 4
        elif code == maxVal and code == c3:
            return 5
        elif code == maxVal and code == c4:
            return 6
        elif code == maxVal and code == c5:
            return 8
        elif code == maxVal and code == c6:
            return 10




#####################################################################################################




def scrape_and_save_to_excel_past(url, output_file):
    # Send a GET request to the specified URL
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    }

    response = requests.get(url, headers=headers)

    
    
    if response.status_code == 200:
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Create a new Excel workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        
        # Find the first <body> element
        body = soup.find('body')
        
        # Find all <tr> elements within the <body>
        rows = body.find_all('tr')
        
        for row in rows:
            # Skip rows that are None
            if row is None:
                continue
            
            #NUMBER search
            rueckennummer_element = row.find('td', class_='rueckennummer')
            number = rueckennummer_element.find('div', class_='rn_nummer').text if rueckennummer_element else None
            #number = int(number) if number and number != '-' else None
            
            # Check if the 'number' is '-' and replace it with a random number
            if number and number != '-' :
                number = int(number)
                used_numbers.append(number)
            if number == '-':
                number = generate_random_number(used_numbers)

            #Image URL search
            posrela_element = row.find('td', class_='posrela')

            img_url_element = posrela_element.find('img') if posrela_element else None
            img_url = img_url_element['data-src'] if img_url_element and 'data-src' in img_url_element.attrs else None

            #NAME search
            name_element = posrela_element.find('a') if posrela_element else None
            name = name_element.text if name_element else None
            
            # Extract the first word as first_name and the rest as last_name
            if name:
                name = name.strip()
                name_parts = name.split(' ', 1)
                first_name = name_parts[0] if name_parts else None
                last_name = name_parts[1] if len(name_parts) > 1 else None
            else:
                first_name = None
                last_name = None
            #POSITION search
            #position_element = positionTable.find_next('tr').find('td') if positionTable else None
            #position = position_element.text if position_element else None
            position_element = posrela_element.find('td', class_='hauptlink').find_next('tr').find('td') if posrela_element else None
            position = position_element.text.strip() if position_element else None
            
            zentriert_tds = row.find_all('td', class_='zentriert')

            # Extract data from the third td element (index 2)
            birth_date = zentriert_tds[1].text.strip() if len(zentriert_tds) > 2 else 'NA'

            zentriert_td = row.find('td', class_='zentriert')
            #birth_date_element = zentrient_td.find_next_siblings('td')[0] if zentrient_td else None
            #birth_date = birth_date_element.text if birth_date_element else None
            
            #NATION search
            nation_element = zentriert_td.find_next_siblings('td')[2].find('img') if zentriert_td else None
            nation = nation_element['alt'] if nation_element else 'NA'
            
            #HEIGHT search
            height_element = zentriert_td.find_next_siblings('td')[4] if zentriert_td else None
            height = height_element.text if height_element else 'NA'
            
            #STRONG_FOOT search
            strong_foot_element = zentriert_td.find_next_siblings('td')[5] if zentriert_td else None
            strong_foot = strong_foot_element.text if strong_foot_element else None
           
            #JOIN search
            join_date_element = zentriert_td.find_next_siblings('td')[6] if zentriert_td else None
            join_date = join_date_element.text if join_date_element and join_date_element.text != '-' else 'NA'
            
            #MARKET value
            value_element = row.find('td', class_='rechts')
            value = value_element.find('a').text if value_element and value_element.find('a') else 'NA'
            
            # Save the extracted data to the Excel sheet
            #ws.append([number, img_url, name, position, birth_date, nation, height, strong_foot, join_date, value])
            #if number and img_url and name and position and birth_date and nation and height and strong_foot and join_date and value:
            #    ws.append([number, img_url, first_name, last_name, position, birth_date, nation, height, strong_foot, join_date, value])
            # Save the extracted data to the Excel sheet
            if number and img_url and first_name and last_name and position and birth_date and nation and height and join_date and value:
                # Check if strong_foot is not 'right' or 'left' and set it to 'right'
                if strong_foot not in ['right', 'left']:
                    strong_foot = 'right'
                if not last_name:
                    last_name = first_name

                ws.append([number, img_url, first_name, last_name, position, birth_date, nation, height, strong_foot, join_date, value])

        
        # Save the workbook to the specified file
        wb.save(output_file)
        print(f"Data has been successfully scraped and saved to {output_file}")
    else:
        print(f"Failed to retrieve data from {url}. Status code: {response.status_code}")











def scrape_and_save_to_excel_current(url, output_file):
    # Send a GET request to the specified URL
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    }

    response = requests.get(url, headers=headers)

    
    
    if response.status_code == 200:
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Create a new Excel workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        
        # Find the first <body> element
        body = soup.find('body')
        
        # Find all <tr> elements within the <body>
        rows = body.find_all('tr')
        
        for row in rows:
            # Skip rows that are None
            if row is None:
                continue
            
            #NUMBER search
            rueckennummer_element = row.find('td', class_='rueckennummer')
            number = rueckennummer_element.find('div', class_='rn_nummer').text if rueckennummer_element else None
            #number = int(number) if number and number != '-' else None
            
            # Check if the 'number' is '-' and replace it with a random number
            if number and number != '-' :
                number = int(number)
                used_numbers.append(number)
            if number == '-':
                number = generate_random_number(used_numbers)

            #Image URL search
            posrela_element = row.find('td', class_='posrela')

            #img_url_element = posrela_element.find('img') if posrela_element else None
            if posrela_element:
                img_url_element = posrela_element.find('img') if posrela_element else None
            else:
                img_url_element = row.find('img')#row.find('table', class_='inline-table').find('img')
            img_url = img_url_element['data-src'] if img_url_element and 'data-src' in img_url_element.attrs else None

            #NAME search
            if posrela_element:
                name_element = posrela_element.find('a') if posrela_element else None
            else:
                name_element = row.find('a')    
            name = name_element.text if name_element else None
            
            # Extract the first word as first_name and the rest as last_name
            if name:
                name = name.strip()
                name_parts = name.split(' ', 1)
                first_name = name_parts[0] if name_parts else None
                last_name = name_parts[1] if len(name_parts) > 1 else None
            else:
                first_name = None
                last_name = None
            #POSITION search
            #position_element = positionTable.find_next('tr').find('td') if positionTable else None
            #position = position_element.text if position_element else None
            position_element = posrela_element.find('td', class_='hauptlink').find_next('tr').find('td') if posrela_element else None
            position = position_element.text.strip() if position_element else None
            
            zentriert_tds = row.find_all('td', class_='zentriert')

            # Extract data from the third td element (index 2)
            birth_date = zentriert_tds[1].text.strip() if len(zentriert_tds) > 2 else 'NA'

            zentriert_td = row.find('td', class_='zentriert')
            #birth_date_element = zentrient_td.find_next_siblings('td')[0] if zentrient_td else None
            #birth_date = birth_date_element.text if birth_date_element else None
            
            #NATION search
            nation_element = zentriert_td.find_next_siblings('td')[2].find('img') if zentriert_td else None
            nation = nation_element['alt'] if nation_element else 'NA'
            
            #HEIGHT search
            height_element = zentriert_td.find_next_siblings('td')[3] if zentriert_td else None
            height = height_element.text if height_element else 'NA'
            
            #STRONG_FOOT search
            strong_foot_element = zentriert_td.find_next_siblings('td')[4] if zentriert_td else None
            strong_foot = strong_foot_element.text if strong_foot_element else None
           
            #JOIN search
            join_date_element = zentriert_td.find_next_siblings('td')[5] if zentriert_td else None
            join_date = join_date_element.text if join_date_element and join_date_element.text != '-' else 'NA'
            
            #CONTRACT search
            contract_date_element = zentriert_td.find_next_siblings('td')[7] if zentriert_td else None
            contract_date = contract_date_element.text if contract_date_element else 'NA'

            #MARKET value
            value_element = row.find('td', class_='rechts')
            value = value_element.find('a').text if value_element and value_element.find('a') else 'NA'
            
            

            # Save the extracted data to the Excel sheet
            #ws.append([number, img_url, name, position, birth_date, nation, height, strong_foot, join_date, value])
            #if number and img_url and name and position and birth_date and nation and height and strong_foot and join_date and value:
            #    ws.append([number, img_url, first_name, last_name, position, birth_date, nation, height, strong_foot, join_date, value])
            # Save the extracted data to the Excel sheet
            if number and img_url and first_name and last_name and position and birth_date and nation and height and join_date and contract_date:
                # Check if strong_foot is not 'right' or 'left' and set it to 'right'
                if strong_foot not in ['right', 'left']:
                    strong_foot = 'right'
                if contract_date in ['-', ' , 0']:
                    contract_date = 'NA'
                if not last_name:
                    last_name = first_name

            if number and position and birth_date and nation and height and join_date and contract_date:
                # Check if strong_foot is not 'right' or 'left' and set it to 'right'
                if strong_foot not in ['right', 'left']:
                    strong_foot = 'right'
                if contract_date in ['-', ' , 0']:
                    contract_date = 'NA'
                if not last_name:
                    last_name = first_name

            if img_url and first_name and last_name:
                # Check if strong_foot is not 'right' or 'left' and set it to 'right'
                if strong_foot not in ['right', 'left']:
                    strong_foot = 'right'
                if contract_date in ['-', ' , 0']:
                    contract_date = 'NA'
                if not last_name:
                    last_name = first_name
            
            if first_name and not last_name:
                last_name = first_name
            
            if height == '-':
                height = 177

            ws.append([number, img_url, first_name, last_name, position, birth_date, nation, height, strong_foot, join_date, contract_date, value])

        
        # Save the workbook to the specified file
        wb.save(output_file)
        print(f"Data has been successfully scraped and saved to {output_file}")
    else:
        print(f"Failed to retrieve data from {url}. Status code: {response.status_code}")
    

    







# Prompt the user to input the URL
squadRecency = input("Is the squad(s) current? (y/n) : ")
url_to_scrape = input("Enter the URL to scrape: ")
output_excel_file = 'output_data.xlsx'

if squadRecency == 'y':
    scrape_and_save_to_excel_current(url_to_scrape, output_excel_file)
    # Load the workbook
    workbook = openpyxl.load_workbook(output_excel_file)

    # Select the active sheet
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
        # Check if the cell in the first column is not empty
        if row[0].value is not None and all(cell.value is None for cell in row[1:]):
            # Get the values of the three cells below and set them in the corresponding cells
            for i in range(1, 4):  # Assuming you want to copy three cells below
                cell_below = sheet.cell(row=row[0].row + 1, column=row[0].column + i)
                cell_to_set = sheet.cell(row=row[0].row, column=row[0].column + i)
                cell_to_set.value = cell_below.value

    # Collect rows to delete in a list
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        # Check if the cell in the first column is not empty
        if row[0].value is None:
            # Add the row to the list of rows to delete
            rows_to_delete.append(row[0].row)

    # Iterate over the list of rows to delete in reverse order
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index)


    # Columns to iterate through
    third_column = sheet['C']
    fourth_column = sheet['D']
    fifth_column = sheet['E']
    sixth_column = sheet['F']
    seventh_column = sheet['G']
    eighth_column = sheet['H']
    ninth_column = sheet['I']
    tenth_column = sheet['J']
    eleventh_column = sheet['K']
    twelfth_column = sheet['L']

    # Iterate through each cell in the third column
    for cell in third_column:
        if cell.value is not None and isinstance(cell.value, str):
            originalFirstN = cell.value
            converted_FirstN = names_Dict.get(originalFirstN)
            if converted_FirstN:
                cell.value = int(converted_FirstN)
            else:
                names_Dict.update({originalFirstN: (int(max(names_Dict.values())) + 1)})
                cell.value = names_Dict.get(originalFirstN)

    # Iterate through each cell in the fourth column
    for cell in fourth_column:
        if cell.value is not None and isinstance(cell.value, str):
            originalLastN = cell.value
            converted_LastN = names_Dict.get(originalLastN)
            if converted_LastN:
                cell.value = int(converted_LastN)
            else:
                names_Dict.update({originalLastN: (int(max(names_Dict.values())) + 1)})
                cell.value = names_Dict.get(originalLastN)


    for cell in fifth_column:
        if cell.value is not None and isinstance(cell.value, str):
            converted_position = cell.value.replace("Goalkeeper", "0").replace("Centre-Back", "5").replace("Left-Back", "7").replace("Right-Back", "3").replace("Defensive Midfield", "10").replace("Mittelfeld", "14").replace("Central Midfield", "14").replace("Attacking Midfield", "18").replace("Left Winger", "27").replace("Right Winger", "23").replace("Centre-Forward", "21").replace("Left Midfield", "16").replace("Right Midfield", "12").replace("Second Striker", "25").replace("Striker", "25")
            cell.value = int(converted_position)

    # Iterate through each cell in the sixth column
    for cell in sixth_column:
        # Process the date string if it is not None and is a string
        if cell.value is not None and isinstance(cell.value, str):
            # Extract the date part from the string
            date_string = cell.value.split('(')[0].strip()

            try:
                # Parse the date string and extract year, month, and day
                date_object = datetime.strptime(date_string, "%b %d, %Y")
                year = date_object.year
                month = date_object.month
                day = date_object.day

                # Update the cell value with the formatted date
                #cleanedDate = f"{year}, {month}, {day}"
                cell.value = dateIDGenerator(year, month, day)

            except ValueError:
                # Handle the case where the date string is not in the expected format
                print(f"Error processing date: {cell.value}")

    # Iterate through each cell in the seventh column
    for cell in seventh_column:
        # Remove whitespace, "m"s, and replace commas with "."
        if cell.value is not None and isinstance(cell.value, str):
            originalNation = cell.value
            converted_Nation = nations_Dict.get(originalNation)
            cell.value = int(converted_Nation)

    # Iterate through each cell in the eighth column
    for cell in eighth_column:
        # Remove whitespace, "m"s, and replace commas with "."
        if cell.value is not None and isinstance(cell.value, str):
            cleaned_value = cell.value.replace(" ", "").replace("m", "").replace(",", "")
            cell.value = int(cleaned_value)

    # Iterate through each cell in the ninth column
    for cell in ninth_column:
        # Remove whitespace, "m"s, and replace commas with "."
        if cell.value is not None and isinstance(cell.value, str):
            sf_value = cell.value.replace("right", "1").replace("left", "2")
            cell.value = int(sf_value)

    # Iterate through each cell in the tenth column
    for cell in tenth_column:
        # Process the date string if it is not None and is a string
        if cell.value is not None and isinstance(cell.value, str):
            
            date_string = cell.value

            try:
                # Parse the date string and extract year, month, and day
                date_object = datetime.strptime(date_string, "%b %d, %Y")
                year = date_object.year
                month = date_object.month
                day = date_object.day

                # Update the cell value with the formatted date
                cell.value = dateIDGenerator(year, month, day)

            except ValueError:
                # Handle the case where the date string is not in the expected format
                cell.value = dateIDGenerator(random.randint(2018,2022), random.randint(1,12), random.randint(1,28))

    # Iterate through each cell in the eleventh column
    for cell in eleventh_column:
        # Process the date string if it is not None and is a string
        if cell.value is not None and isinstance(cell.value, str):
            
            date_string = cell.value

            try:
                # Parse the date string and extract year, month, and day
                date_object = datetime.strptime(date_string, "%b %d, %Y")
                year = date_object.year
                month = date_object.month
                day = date_object.day

                # Update the cell value with the formatted date
                cell.value = dateIDGenerator(year, month, day)

            except ValueError:
                # Handle the case where the date string is not in the expected format
                cell.value = dateIDGenerator(random.randint(2018,2022), random.randint(1,12), random.randint(1,28))

    # Iterate through each cell in the twelfth column
    for cell in twelfth_column:
        if cell.value is not None and isinstance(cell.value, str) and 'k' not in cell.value:
            cleaned_value = cell.value.replace("€", "").replace("m", "")
            cell.value = float(cleaned_value)
        else:
            cleaned_value = cell.value.replace("€", "").replace("m", "").replace("k", "")
            cleaned_value = float(cleaned_value) / 1000
            cell.value = float(cleaned_value)


    # Save the modified workbook
    workbook.save(output_excel_file)


    # Convert the dictionary to a DataFrame
    finalNames = pd.DataFrame(list(names_Dict.items()), columns=['name', 'nameid'])
    # Specify the CSV file path
    dc_file_path = 'dcplayernames.txt'
    # Write the DataFrame to a text file with utf-16-le encoding and tab separator
    with open(dc_file_path, 'w', encoding='utf-16-le') as file:
        file.write('\t'.join(finalNames.columns) + '\n')  # Write header
        for _, row in finalNames.iterrows():
            file.write('\t'.join(map(str, row)) + '\n')  # Write each row








elif squadRecency == 'n':
    scrape_and_save_to_excel_past(url_to_scrape, output_excel_file)

    # Load the workbook
    workbook = openpyxl.load_workbook(output_excel_file)

    # Select the active sheet
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
        # Check if the cell in the first column is not empty
        if row[0].value is not None and all(cell.value is None for cell in row[1:]):
            # Get the values of the three cells below and set them in the corresponding cells
            for i in range(1, 4):  # Assuming you want to copy three cells below
                cell_below = sheet.cell(row=row[0].row + 1, column=row[0].column + i)
                cell_to_set = sheet.cell(row=row[0].row, column=row[0].column + i)
                cell_to_set.value = cell_below.value

    # Collect rows to delete in a list
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        # Check if the cell in the first column is not empty
        if row[0].value is None:
            # Add the row to the list of rows to delete
            rows_to_delete.append(row[0].row)

    # Iterate over the list of rows to delete in reverse order
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index)


    # Columns to iterate through
    third_column = sheet['C']
    fourth_column = sheet['D']
    fifth_column = sheet['E']
    sixth_column = sheet['F']
    seventh_column = sheet['G']
    eighth_column = sheet['H']
    ninth_column = sheet['I']
    tenth_column = sheet['J']

    # Iterate through each cell in the third column
    for cell in third_column:
        if cell.value is not None and isinstance(cell.value, str):
            originalFirstN = cell.value
            converted_FirstN = names_Dict.get(originalFirstN)
            if converted_FirstN:
                cell.value = int(converted_FirstN)
            else:
                names_Dict.update({originalFirstN: (int(max(names_Dict.values())) + 1)})
                cell.value = names_Dict.get(originalFirstN)

    # Iterate through each cell in the third column
    for cell in fourth_column:
        if cell.value is not None and isinstance(cell.value, str):
            originalLastN = cell.value
            converted_LastN = names_Dict.get(originalLastN)
            if converted_LastN:
                cell.value = int(converted_LastN)
            else:
                names_Dict.update({originalLastN: (int(max(names_Dict.values())) + 1)})
                cell.value = names_Dict.get(originalLastN)


    for cell in fifth_column:
        if cell.value is not None and isinstance(cell.value, str):
            converted_position = cell.value.replace("Goalkeeper", "0").replace("Centre-Back", "5").replace("Left-Back", "7").replace("Right-Back", "3").replace("Defensive Midfield", "10").replace("Mittelfeld", "14").replace("Central Midfield", "14").replace("Attacking Midfield", "18").replace("Left Winger", "27").replace("Right Winger", "23").replace("Centre-Forward", "21").replace("Left Midfield", "16").replace("Right Midfield", "12").replace("Second Striker", "25").replace("Striker", "25")
            cell.value = int(converted_position)

    # Iterate through each cell in the sixth column
    for cell in sixth_column:
        # Process the date string if it is not None and is a string
        if cell.value is not None and isinstance(cell.value, str):
            # Extract the date part from the string
            date_string = cell.value.split('(')[0].strip()

            try:
                # Parse the date string and extract year, month, and day
                date_object = datetime.strptime(date_string, "%b %d, %Y")
                year = date_object.year
                month = date_object.month
                day = date_object.day

                # Update the cell value with the formatted date
                #cleanedDate = f"{year}, {month}, {day}"
                cell.value = dateIDGenerator(year, month, day)

            except ValueError:
                # Handle the case where the date string is not in the expected format
                print(f"Error processing date: {cell.value}")

    # Iterate through each cell in the seventh column
    for cell in seventh_column:
        # Remove whitespace, "m"s, and replace commas with "."
        if cell.value is not None and isinstance(cell.value, str):
            originalNation = cell.value
            converted_Nation = nations_Dict.get(originalNation)
            cell.value = int(converted_Nation)

    # Iterate through each cell in the eighth column
    for cell in eighth_column:
        # Remove whitespace, "m"s, and replace commas with "."
        if cell.value is not None and isinstance(cell.value, str):
            cleaned_value = cell.value.replace(" ", "").replace("m", "").replace(",", "")
            cell.value = int(cleaned_value)

    # Iterate through each cell in the ninth column
    for cell in ninth_column:
        # Remove whitespace, "m"s, and replace commas with "."
        if cell.value is not None and isinstance(cell.value, str):
            sf_value = cell.value.replace("right", "1").replace("left", "2")
            cell.value = int(sf_value)

    # Iterate through each cell in the tenth column
    for cell in tenth_column:
        # Process the date string if it is not None and is a string
        if cell.value is not None and isinstance(cell.value, str):
            
            date_string = cell.value

            try:
                # Parse the date string and extract year, month, and day
                date_object = datetime.strptime(date_string, "%b %d, %Y")
                year = date_object.year
                month = date_object.month
                day = date_object.day

                # Update the cell value with the formatted date
                cell.value = dateIDGenerator(year, month, day)

            except ValueError:
                # Handle the case where the date string is not in the expected format
                cell.value = dateIDGenerator(random.randint(2018,2022), random.randint(1,12), random.randint(1,28))


    # Save the modified workbook
    workbook.save(output_excel_file)


    # Convert the dictionary to a DataFrame
    finalNames = pd.DataFrame(list(names_Dict.items()), columns=['name', 'nameid'])
    # Specify the CSV file path
    dc_file_path = 'dcplayernames.txt'
    
    # Write the DataFrame to a text file with utf-16-le encoding and tab separator
    with open(dc_file_path, 'w', encoding='utf-16-le') as file:
        file.write('\t'.join(finalNames.columns) + '\n')  # Write header
        for _, row in finalNames.iterrows():
            file.write('\t'.join(map(str, row)) + '\n')  # Write each row


else:
    print("Errored input, please try running the script again")







#########################################################################################################


# Player data converted into rows of data to be appended to the players table


#########################################################################################################





defaultPlayerTable = 'players.txt'
newSquad = 'newSquad.txt'



def playerTableHandler(playerTable, excelFile):
    # Load the workbook from the output file
    output_workbook = openpyxl.load_workbook(output_excel_file)
    output_sheet = output_workbook.active

    # Load the workbook from the players file
    #players_workbook = openpyxl.load_workbook(defaultPlayerTable)
    #players_sheet = players_workbook.active

    start_row_number = 20249

    playerIDCounter = 80000

    # Open the text file in write mode
    with open(newSquad, 'w', encoding='utf-16-le') as new_text_file:
        # Iterate through rows in the Excel file
        for row_number, row in enumerate(output_sheet.iter_rows(min_row=0, max_row=output_sheet.max_row, min_col=1, max_col=127), start=start_row_number):
        # Example: Extract data from Excel columns and store in variables
                birthdate = row[5].value
                overallrating = attackerEstimation(ageGetter(birthdate), row[11].value, 300)

                firstnameid = row[2].value
                lastnameid = row[3].value
                playerjerseynameid = lastnameid
                commonnameid = lastnameid
                skintypecode = 0
                trait2 = 0
                haircolorcode = 1
                facialhairtypecode = predictFacialHair(row[1].value)
                curve = random.randint(overallrating-5, overallrating+5)
                jerseystylecode = 1
                agility = random.randint(overallrating-5, overallrating+5)
                tattooback = 0
                accessorycode4 = 0
                gksavetype = 0
                positioning = random.randint(overallrating-5, overallrating+5)
                tattooleftarm = 0
                hairtypecode = predictHairstyle(row[1].value)
                standingtackle = random.randint(overallrating-5, overallrating+5)
                preferredposition3 = -1
                longpassing = random.randint(overallrating-5, overallrating+5)
                penalties = random.randint(overallrating-5, overallrating+5)
                animfreekickstartposcode = 0
                isretiring = 0
                longshots = random.randint(overallrating-5, overallrating+5)
                gkdiving = random.randint(overallrating-5, overallrating+5)
                interceptions = random.randint(overallrating-5, overallrating+5)
                shoecolorcode2 = 23
                crossing = random.randint(overallrating-5, overallrating+5)
                potential = random.randint(overallrating, overallrating+5)
                gkreflexes = random.randint(overallrating-5, overallrating+5)
                finishingcode1 = 0
                reactions = random.randint(overallrating-5, overallrating+5)
                composure = random.randint(overallrating-5, overallrating+5)
                vision = random.randint(overallrating-5, overallrating+5)
                contractvaliduntil = 2025
                finishing = random.randint(overallrating-5, overallrating+5)
                dribbling = random.randint(overallrating-5, overallrating+5)
                slidingtackle = random.randint(overallrating-5, overallrating+5)
                accessorycode3 = 0
                accessorycolourcode1 = 0
                headtypecode = 0
                driref = random.randint(overallrating-5, overallrating+5)
                sprintspeed = random.randint(overallrating-5, overallrating+5)
                height = row[7].value
                hasseasonaljersey = 0
                tattoohead = 0
                preferredposition2 = -1
                strength = random.randint(overallrating-5, overallrating+5)
                shoetypecode = 42
                birthdate = row[5].value
                preferredposition1 = row[4].value
                tattooleftleg = 0
                ballcontrol = random.randint(overallrating-5, overallrating+5)
                phypos = random.randint(overallrating-5, overallrating+5)
                shotpower = random.randint(overallrating-5, overallrating+5)
                trait1 = 0
                socklengthcode = 2
                weight = 70
                hashighqualityhead = 0
                gkglovetypecode = 0
                tattoorightarm = 0
                balance = random.randint(overallrating-5, overallrating+5)
                gender = 0
                headassetid = 0
                gkkicking = random.randint(overallrating-5, overallrating+5)
                defspe = random.randint(overallrating-5, overallrating+5)
                internationalrep = 1
                shortpassing = random.randint(overallrating-5, overallrating+5)
                freekickaccuracy = random.randint(overallrating-5, overallrating+5)
                skillmoves = 3
                faceposerpreset = 0
                usercaneditname = 1
                avatarpomid = 0
                attackingworkrate = 2
                finishingcode2 = 0
                aggression = random.randint(overallrating-5, overallrating+5)
                acceleration = random.randint(overallrating-5, overallrating+5)
                paskic = random.randint(overallrating-5, overallrating+5)
                headingaccuracy = random.randint(overallrating-5, overallrating+5)
                iscustomized = 0
                eyebrowcode = 0
                runningcode2 = 0
                modifier = 0
                gkhandling = random.randint(overallrating-5, overallrating+5)
                eyecolorcode = 5
                jerseysleevelengthcode = 0
                accessorycolourcode3 = 0
                accessorycode1 = 0
                playerjointeamdate = row[9].value
                headclasscode = 1
                defensiveworkrate = 2
                tattoofront = 0
                nationality = row[6].value
                preferredfoot = row[8].value
                sideburnscode = 0
                weakfootabilitytypecode = 3
                jumping = random.randint(overallrating-5, overallrating+5)
                personality = 2
                gkkickstyle = 0
                stamina = random.randint(overallrating-5, overallrating+5)

                
                playerid = playerIDCounter
                playerIDCounter += 1

                marking = random.randint(overallrating-5, overallrating+5)
                accessorycolourcode4 = 0
                gkpositioning = random.randint(overallrating-5, overallrating+5)
                headvariation = 0
                skillmoveslikelihood = 3
                shohan = random.randint(overallrating-5, overallrating+5)
                skintonecode = predictSkinTone(row[1].value)
                shortstyle = 0

                #overallrating = attackerEstimation(ageGetter(birthdate), row[11].value, 55)

                smallsidedshoetypecode = 500
                emotion = 1
                runstylecode = 0
                jerseyfit = 0
                accessorycode2 = 0
                shoedesigncode = 0
                shoecolorcode1 = 0
                hairstylecode = 0
                bodytypecode = 2
                animpenaltiesstartposcode = 0
                pacdiv = random.randint(overallrating-5, overallrating+5)
                runningcode1 = 0
                preferredposition4 = -1
                volleys = random.randint(overallrating-5, overallrating+5)
                accessorycolourcode2 = 0
                tattoorightleg = 0
                facialhaircolorcode = 0


                # Construct a row with the modified data
                row_data = [
                    firstnameid,
                    lastnameid,
                    playerjerseynameid,
                    commonnameid,
                    skintypecode,
                    trait2,
                    haircolorcode,
                    facialhairtypecode,
                    curve,
                    jerseystylecode,
                    agility,
                    tattooback,
                    accessorycode4,
                    gksavetype,
                    positioning,
                    tattooleftarm,
                    hairtypecode,
                    standingtackle,
                    preferredposition3,
                    longpassing,
                    penalties,
                    animfreekickstartposcode,
                    isretiring,
                    longshots,
                    gkdiving,
                    interceptions,
                    shoecolorcode2,
                    crossing,
                    potential,
                    gkreflexes,
                    finishingcode1,
                    reactions,
                    composure,
                    vision,
                    contractvaliduntil,
                    finishing,
                    dribbling,
                    slidingtackle,
                    accessorycode3,
                    accessorycolourcode1,
                    headtypecode,
                    driref,
                    sprintspeed,
                    height,
                    hasseasonaljersey,
                    tattoohead,
                    preferredposition2,
                    strength,
                    shoetypecode,
                    birthdate,
                    preferredposition1,
                    tattooleftleg,
                    ballcontrol,
                    phypos,
                    shotpower,
                    trait1,
                    socklengthcode,
                    weight,
                    hashighqualityhead,
                    gkglovetypecode,
                    tattoorightarm,
                    balance,
                    gender,
                    headassetid,
                    gkkicking,
                    defspe,
                    internationalrep,
                    shortpassing,
                    freekickaccuracy,
                    skillmoves,
                    faceposerpreset,
                    usercaneditname,
                    avatarpomid,
                    attackingworkrate,
                    finishingcode2,
                    aggression,
                    acceleration,
                    paskic,
                    headingaccuracy,
                    iscustomized,
                    eyebrowcode,
                    runningcode2,
                    modifier,
                    gkhandling,
                    eyecolorcode,
                    jerseysleevelengthcode,
                    accessorycolourcode3,
                    accessorycode1,
                    playerjointeamdate,
                    headclasscode,
                    defensiveworkrate,
                    tattoofront,
                    nationality,
                    preferredfoot,
                    sideburnscode,
                    weakfootabilitytypecode,
                    jumping,
                    personality,
                    gkkickstyle,
                    stamina,
                    playerid,
                    marking,
                    accessorycolourcode4,
                    gkpositioning,
                    headvariation,
                    skillmoveslikelihood,
                    shohan,
                    skintonecode,
                    shortstyle,
                    overallrating,
                    smallsidedshoetypecode,
                    emotion,
                    runstylecode,
                    jerseyfit,
                    accessorycode2,
                    shoedesigncode,
                    shoecolorcode1,
                    hairstylecode,
                    bodytypecode,
                    animpenaltiesstartposcode,
                    pacdiv,
                    runningcode1,
                    preferredposition4,
                    volleys,
                    accessorycolourcode2,
                    tattoorightleg,
                    facialhaircolorcode
                ]


                # Convert the row data to a tab-delimited string
                row_string = '\t'.join(map(str, row_data))

                # Write the row string to the text file
                new_text_file.write(row_string + '\n')
 

    # Save the changes to the players Excel file
    #players_workbook.save(defaultPlayerTable)
    print('players.txt has been modified and saved as combined.txt')


















playerTableHandler(defaultPlayerTable, output_excel_file)




# Reading the original file
with open(defaultPlayerTable, 'r', encoding='utf-16-le') as original_file:
    original_data = original_file.read().strip()

# Reading the new file
with open(newSquad, 'r', encoding='utf-16-le') as new_file:
    new_data = new_file.read().strip()

# Combining data and writing to another file
combined_data = original_data + '\n' + new_data
with open('combined.txt', 'w', encoding='utf-16-le') as combined_file:
    combined_file.write(combined_data)





def playerTeamLinker():
    # Load the workbook from the output file
    output_workbook = openpyxl.load_workbook(output_excel_file)
    output_sheet = output_workbook.active

    start_row_number = 20249

    playerIDCounter = 80000
    artificalKeyCounter = 21023

    newPlayerLinks = 'newPlayerLinks.txt'

    # Open the text file in write mode
    with open(newPlayerLinks, 'w', encoding='utf-16-le') as new_text_file:
        # Iterate through rows in the Excel file
        for row_number, row in enumerate(output_sheet.iter_rows(min_row=0, max_row=output_sheet.max_row, min_col=1, max_col=127), start=start_row_number):
        # Example: Extract data from Excel columns and store in variables
            leaguegoals = 0
            isamongtopscorers = 0
            yellows = 0
            isamongtopscorersinteam = 0
            jerseynumber = row[0].value
            position = row[4].value

            artificialkey = artificalKeyCounter
            artificalKeyCounter += 1
            
            teamid = 114899 #MAZATLAN
            leaguegoalsprevmatch = 0
            injury = 0
            leagueappearances = 0
            istopscorer = 0
            leaguegoalsprevthreematches = 0

            
            playerid = playerIDCounter
            playerIDCounter += 1

            form = 0
            reds = 0
            # Construct a row with the modified data
            row_data = [
                leaguegoals,
                isamongtopscorers,
                yellows,
                isamongtopscorersinteam,
                jerseynumber,
                position,
                artificialkey,
                teamid,
                leaguegoalsprevmatch,
                injury,
                leagueappearances,
                istopscorer,
                leaguegoalsprevthreematches,
                playerid,
                form,
                reds
                ]
            
            # Convert the row data to a tab-delimited string
            row_string = '\t'.join(map(str, row_data))

            # Write the row string to the text file
            new_text_file.write(row_string + '\n')


playerTeamLinker()




# Reading the original file
with open('teamplayerlinks.txt', 'r', encoding='utf-16-le') as original_file:
    original_data = original_file.read().strip()

# Reading the new file
with open('newPlayerLinks.txt', 'r', encoding='utf-16-le') as new_file:
    new_data = new_file.read().strip()

# Combining data and writing to another file
combined_data = original_data + '\n' + new_data
with open('newteamplayerlinks.txt', 'w', encoding='utf-16-le') as combined_file:
    combined_file.write(combined_data)